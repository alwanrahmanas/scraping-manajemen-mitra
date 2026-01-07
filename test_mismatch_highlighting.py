"""
Test script untuk demonstrasi fitur mismatch highlighting
Membuat sample Excel dengan data normal dan data mismatch
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def create_sample_excel():
    """Create sample Excel with mismatch highlighting demo"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Mitra"
    
    # Headers
    headers = [
        "NIK", 
        "Nama Lengkap (dengan Gelar)",
        "Nomor Rekening",
        "Nama Bank",
        "Nama Pemilik Rekening",
        "Status"
    ]
    ws.append(headers)
    
    # Header formatting
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sample data - mix of normal and mismatch
    sample_data = [
        # Normal rows
        {
            "NIK": "1234567890123456",
            "Nama": "John Doe, S.Kom",
            "Rekening": "1234567890",
            "Bank": "BRI",
            "Pemilik": "JOHN DOE",
            "mismatch": False
        },
        {
            "NIK": "2345678901234567",
            "Nama": "Jane Smith, S.E.",
            "Rekening": "9876543210",
            "Bank": "BCA",
            "Pemilik": "JANE SMITH",
            "mismatch": False
        },
        # Mismatch rows (nomor rekening contains text)
        {
            "NIK": "3456789012345678",
            "Nama": "Bob Wilson, S.T.",
            "Rekening": "JOHN DOE",  # ← MISMATCH! Ini harusnya angka
            "Bank": "BNI",
            "Pemilik": "1234567890",  # ← Data tertukar!
            "mismatch": True
        },
        {
            "NIK": "4567890123456789",
            "Nama": "Alice Brown, A.Md",
            "Rekening": "5555666677",
            "Bank": "Mandiri",
            "Pemilik": "ALICE BROWN",
            "mismatch": False
        },
        {
            "NIK": "5678901234567890",
            "Nama": "Charlie Davis, S.Sos",
            "Rekening": "BCA SYARIAH",  # ← MISMATCH! Ini harusnya angka
            "Bank": "8888999900",  # ← Data tertukar!
            "Pemilik": "CHARLIE DAVIS",
            "mismatch": True
        },
    ]
    
    # Add data rows with highlighting
    mismatch_count = 0
    for row_idx, data in enumerate(sample_data, start=2):
        ws.append([
            data["NIK"],
            data["Nama"],
            data["Rekening"],
            data["Bank"],
            data["Pemilik"],
            "Success"
        ])
        
        # Highlight mismatch rows
        if data["mismatch"]:
            mismatch_count += 1
            mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            mismatch_font = Font(color="9C0006", bold=True)
            
            for cell in ws[row_idx]:
                cell.fill = mismatch_fill
                # Highlight Nomor Rekening column specifically
                if cell.column == 3:  # Column C
                    cell.font = mismatch_font
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create Summary Sheet
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.append(["SCRAPING SUMMARY & DATA QUALITY REPORT"])
    ws_summary.append([])
    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Total Rows Scraped", len(sample_data)])
    ws_summary.append(["Rows with Potential Mismatch", mismatch_count])
    ws_summary.append(["Data Quality Rate", f"{((len(sample_data) - mismatch_count) / len(sample_data) * 100):.1f}%"])
    ws_summary.append([])
    ws_summary.append(["LEGEND:"])
    ws_summary.append(["🔴 Red/Pink Rows", "= Potential mismatch detected (Nomor Rekening contains non-numeric characters)"])
    ws_summary.append(["⚠️ Action Required", "= Please verify these rows manually"])
    ws_summary.append([])
    ws_summary.append(["Note:", "Mismatch detection helps identify data quality issues where account number may have been incorrectly scraped."])
    ws_summary.append([])
    ws_summary.append(["EXAMPLE MISMATCHES IN THIS FILE:"])
    ws_summary.append(["Row 4 (NIK 3456789012345678)", "Nomor Rekening = 'JOHN DOE' (should be numbers)"])
    ws_summary.append(["Row 6 (NIK 5678901234567890)", "Nomor Rekening = 'BCA SYARIAH' (should be numbers)"])
    
    # Format summary sheet
    ws_summary['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_summary['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    ws_summary.merge_cells('A1:B1')
    
    # Format metric headers
    for cell in ws_summary[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Highlight mismatch count
    if mismatch_count > 0:
        ws_summary['B5'].font = Font(bold=True, color="9C0006")
        ws_summary['B5'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Auto-adjust summary column widths
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 70
    
    # Save file
    output_file = "sample_mismatch_highlighting.xlsx"
    wb.save(output_file)
    
    print("=" * 60)
    print("SAMPLE EXCEL CREATED SUCCESSFULLY!")
    print("=" * 60)
    print(f"File: {output_file}")
    print(f"Total rows: {len(sample_data)}")
    print(f"Mismatch rows: {mismatch_count} (highlighted in RED)")
    print(f"Data quality: {((len(sample_data) - mismatch_count) / len(sample_data) * 100):.1f}%")
    print()
    print("Open the file to see:")
    print("  1. Summary sheet with statistics")
    print("  2. Data Mitra sheet with highlighted mismatch rows")
    print("  3. Red/pink highlighting on rows with invalid account numbers")
    print()
    print("WARNING: Check rows 4 and 6 - they have mismatch!")
    print("=" * 60)

if __name__ == "__main__":
    create_sample_excel()
