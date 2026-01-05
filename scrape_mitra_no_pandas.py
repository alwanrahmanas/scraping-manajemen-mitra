import os
import sys
import logging
import requests
import csv
from datetime import datetime
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Setup logging
log_filename = f"scraper_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

class MitraScraper:
    def __init__(self):
        self.base_download_dir = "downloads"
        self.data_list = []
        self.stats = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'ktp_downloaded': 0,
            'ijazah_downloaded': 0,
            'pages_processed': 0
        }
        
        # Create downloads directory
        if not os.path.exists(self.base_download_dir):
            os.makedirs(self.base_download_dir)
            logger.info(f"Created downloads directory: {self.base_download_dir}")

    def download_image(self, url, folder, filename):
        """Download image from URL with detailed logging"""
        if not url:
            logger.warning(f"No URL provided for {filename}")
            return None
        
        try:
            logger.info(f"Downloading {filename} from {url[:100]}...")
            response = requests.get(url, timeout=15)
            
            if response.status_code == 200:
                path = os.path.join(folder, filename)
                with open(path, 'wb') as f:
                    f.write(response.content)
                
                file_size = len(response.content) / 1024  # KB
                logger.info(f"✓ Downloaded {filename} ({file_size:.2f} KB) -> {path}")
                return path
            else:
                logger.error(f"✗ Failed to download {filename}: HTTP {response.status_code}")
                return None
                
        except Exception as e:
            logger.error(f"✗ Error downloading {filename}: {str(e)}")
            return None

    def extract_bank_info(self, page):
        """Extract bank information from Rekening tab"""
        logger.info("Extracting bank information...")
        
        nama_bank = "N/A"
        no_rekening = "N/A"
        nama_pemilik = "N/A"

        try:
            # Strategy 1: Direct extraction from form-control-plaintext
            # Based on HTML: <label>Nama Bank</label> <div class="form-control-plaintext"><span><span>VALUE</span></span></div>
            
            try:
                # Find the label "Nama Bank" then get the next div with class form-control-plaintext
                bank_container = page.locator('label:has-text("Nama Bank") + div.form-control-plaintext')
                if bank_container.count() > 0:
                    nama_bank = bank_container.inner_text().strip()
                    logger.info(f"Found Nama Bank: {nama_bank}")
            except Exception as e:
                logger.debug(f"Strategy 1 failed for Nama Bank: {e}")
            
            try:
                rek_container = page.locator('label:has-text("Nomor Rekening") + div.form-control-plaintext')
                if rek_container.count() > 0:
                    no_rekening = rek_container.inner_text().strip()
                    logger.info(f"Found Nomor Rekening: {no_rekening}")
            except Exception as e:
                logger.debug(f"Strategy 1 failed for Nomor Rekening: {e}")
            
            try:
                owner_container = page.locator('label:has-text("Nama Pemilik Rekening") + div.form-control-plaintext')
                if owner_container.count() > 0:
                    nama_pemilik = owner_container.inner_text().strip()
                    logger.info(f"Found Nama Pemilik: {nama_pemilik}")
            except Exception as e:
                logger.debug(f"Strategy 1 failed for Nama Pemilik: {e}")
            
            # Fallback: Parse from text dump if any field is still N/A
            if nama_bank == "N/A" or no_rekening == "N/A" or nama_pemilik == "N/A":
                logger.info("Using fallback text parsing...")
                modal_text = page.locator(".v--modal-box").inner_text()
                lines = modal_text.split('\n')
                
                for i, line in enumerate(lines):
                    line = line.strip()
                    
                    if nama_bank == "N/A" and "Nama Bank" in line and i + 1 < len(lines):
                        potential_bank = lines[i + 1].strip()
                        if potential_bank and ("BANK" in potential_bank.upper() or potential_bank.startswith("(")):
                            nama_bank = potential_bank
                            logger.info(f"Fallback found Nama Bank: {nama_bank}")
                    
                    if no_rekening == "N/A" and "Nomor Rekening" in line and i + 1 < len(lines):
                        potential_rek = lines[i + 1].strip()
                        if potential_rek and (potential_rek.replace(' ', '').isdigit() or len(potential_rek) > 8):
                            no_rekening = potential_rek
                            logger.info(f"Fallback found Nomor Rekening: {no_rekening}")
                    
                    if nama_pemilik == "N/A" and "Nama Pemilik" in line and i + 1 < len(lines):
                        potential_owner = lines[i + 1].strip()
                        if potential_owner and len(potential_owner) > 2 and not potential_owner.isdigit():
                            nama_pemilik = potential_owner
                            logger.info(f"Fallback found Nama Pemilik: {nama_pemilik}")
            
        except Exception as e:
            logger.error(f"Error extracting bank info: {str(e)}")
        
        return nama_bank, no_rekening, nama_pemilik

    def process_row(self, row, index, page):
        """Process a single table row"""
        try:
            # Find NIK link
            nik_link = row.locator('span[title="Lihat Detail Mitra"]')
            if not nik_link.count():
                logger.debug(f"Row {index}: No NIK link found, skipping")
                return False
            
            nik_text = nik_link.inner_text().strip()
            logger.info(f"\n{'='*60}")
            logger.info(f"Processing Row {index + 1}: NIK {nik_text}")
            logger.info(f"{'='*60}")
            
            # Create user directory
            user_download_dir = os.path.join(self.base_download_dir, nik_text)
            if not os.path.exists(user_download_dir):
                os.makedirs(user_download_dir)
                logger.info(f"Created directory: {user_download_dir}")
            
            # Click NIK to open popup
            logger.info("Opening detail popup...")
            nik_link.click()
            page.wait_for_selector("text=Detail Informasi Mitra", timeout=5000)
            logger.info("✓ Popup opened")
            
            # === Tab 1: File Administrasi ===
            # Note: Default tab is "Profil", so we need to click "File Administrasi"
            logger.info("\n--- Processing File Administrasi ---")
            try:
                page.locator('.nav-link:has-text("File Administrasi")').click(timeout=5000)
                logger.info("✓ Clicked File Administrasi tab")
            except Exception as e:
                logger.warning(f"Could not click File Administrasi tab: {e}")
                try:
                    page.locator('[role="tab"]:has-text("File Administrasi")').click(timeout=5000)
                    logger.info("✓ Clicked File Administrasi tab (fallback)")
                except:
                    logger.error("Failed to click File Administrasi tab")
            
            page.wait_for_timeout(1000)  # Wait for content to load
            
            ktp_path = None
            ijazah_path = None
            
            # Find images
            images = page.locator(".v--modal-box img").all()
            logger.info(f"Found {len(images)} images in modal")
            
            if len(images) >= 1:
                ktp_src = images[0].get_attribute("src")
                ktp_path = self.download_image(ktp_src, user_download_dir, "ktp.jpg")
                if ktp_path:
                    self.stats['ktp_downloaded'] += 1
            
            if len(images) >= 2:
                ijazah_src = images[1].get_attribute("src")
                ijazah_path = self.download_image(ijazah_src, user_download_dir, "ijazah.jpg")
                if ijazah_path:
                    self.stats['ijazah_downloaded'] += 1
            
            # === Tab 2: Rekening ===
            logger.info("\n--- Processing Rekening ---")
            try:
                page.locator('.nav-link:has-text("Rekening")').click(timeout=5000)
                logger.info("✓ Clicked Rekening tab")
            except Exception as e:
                logger.warning(f"Could not click Rekening tab with nav-link: {e}")
                try:
                    page.locator('[role="tab"]:has-text("Rekening")').click(timeout=5000)
                    logger.info("✓ Clicked Rekening tab (fallback)")
                except Exception as e2:
                    logger.error(f"Failed to click Rekening tab: {e2}")
            
            page.wait_for_timeout(1000)
            
            nama_bank, no_rekening, nama_pemilik = self.extract_bank_info(page)
            
            # Store data
            row_data = {
                "NIK": nik_text,
                "Nama Bank": nama_bank,
                "Nomor Rekening": no_rekening,
                "Nama Pemilik": nama_pemilik,
                "Path KTP": ktp_path if ktp_path else "Not Downloaded",
                "Path Ijazah": ijazah_path if ijazah_path else "Not Downloaded",
                "Status": "Success"
            }
            self.data_list.append(row_data)
            
            logger.info(f"\n✓ Successfully processed NIK {nik_text}")
            logger.info(f"  Bank: {nama_bank}")
            logger.info(f"  Rekening: {no_rekening}")
            logger.info(f"  Pemilik: {nama_pemilik}")
            
            # Close modal
            page.keyboard.press("Escape")
            page.wait_for_timeout(500)
            
            self.stats['success'] += 1
            return True
            
        except Exception as e:
            logger.error(f"✗ Error processing row {index}: {str(e)}", exc_info=True)
            self.stats['failed'] += 1
            
            # Try to close modal and recover
            try:
                page.keyboard.press("Escape")
                page.wait_for_timeout(500)
            except:
                pass
            
            # Store failed entry
            self.data_list.append({
                "NIK": "Unknown",
                "Nama Bank": "N/A",
                "Nomor Rekening": "N/A",
                "Nama Pemilik": "N/A",
                "Path KTP": "Failed",
                "Path Ijazah": "Failed",
                "Status": f"Failed: {str(e)[:100]}"
            })
            
            return False

    def save_to_excel(self, filename="mitra_data.xlsx"):
        """Save data to Excel with formatting (without pandas)"""
        logger.info(f"\nSaving data to Excel: {filename}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Mitra"
        
        # Headers
        headers = ["NIK", "Nama Bank", "Nomor Rekening", "Nama Pemilik", "Path KTP", "Path Ijazah", "Status"]
        ws.append(headers)
        
        # Header formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data rows
        for row_data in self.data_list:
            ws.append([
                row_data.get("NIK", ""),
                row_data.get("Nama Bank", ""),
                row_data.get("Nomor Rekening", ""),
                row_data.get("Nama Pemilik", ""),
                row_data.get("Path KTP", ""),
                row_data.get("Path Ijazah", ""),
                row_data.get("Status", "")
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        
        wb.save(filename)
        logger.info(f"✓ Excel file saved with formatting: {filename}")

    def save_to_csv(self, filename="mitra_data.csv"):
        """Save data to CSV"""
        logger.info(f"Saving CSV backup: {filename}")
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=["NIK", "Nama Bank", "Nomor Rekening", "Nama Pemilik", "Path KTP", "Path Ijazah", "Status"])
            writer.writeheader()
            writer.writerows(self.data_list)
        
        logger.info(f"✓ CSV backup saved: {filename}")

    def print_summary(self):
        """Print scraping summary"""
        logger.info(f"\n{'='*60}")
        logger.info("SCRAPING SUMMARY")
        logger.info(f"{'='*60}")
        logger.info(f"Pages processed: {self.stats['pages_processed']}")
        logger.info(f"Total rows processed: {self.stats['total']}")
        logger.info(f"✓ Successful: {self.stats['success']}")
        logger.info(f"✗ Failed: {self.stats['failed']}")
        logger.info(f"📷 KTP downloaded: {self.stats['ktp_downloaded']}")
        logger.info(f"📷 Ijazah downloaded: {self.stats['ijazah_downloaded']}")
        logger.info(f"{'='*60}")

    def run(self):
        """Main scraping process"""
        logger.info("="*60)
        logger.info("MITRA BPS SCRAPER - STARTING")
        logger.info("="*60)
        logger.info(f"Log file: {log_filename}")
        
        with sync_playwright() as p:
            try:
                # Connect to browser
                logger.info("\nConnecting to Chrome (port 9222)...")
                browser = p.chromium.connect_over_cdp("http://localhost:9222")
                context = browser.contexts[0]
                page = context.pages[0]
                
                logger.info(f"✓ Connected to: {page.title()}")
                logger.info(f"✓ URL: {page.url}")
                
                # Wait for table with multiple strategies
                logger.info("\nWaiting for table to load...")
                
                # Strategy 1: Wait for table to exist (not necessarily visible)
                try:
                    page.wait_for_selector("table#vgt-table tbody tr", state="attached", timeout=5000)
                    logger.info("✓ Table found (attached)")
                except:
                    logger.warning("Table not found with 'attached' state, trying alternative...")
                
                # Strategy 2: Wait for loading overlay to disappear
                try:
                    # Check if there's a loading overlay
                    overlay = page.locator(".velmld-overlay")
                    if overlay.count() > 0:
                        logger.info("Waiting for loading overlay to disappear...")
                        page.wait_for_selector(".velmld-overlay", state="hidden", timeout=15000)
                        logger.info("✓ Loading overlay hidden")
                except:
                    logger.info("No loading overlay found or already hidden")
                
                # Strategy 3: Just wait a bit for any animations
                page.wait_for_timeout(2000)
                
                # Get all rows that have NIK links (skip header rows)
                logger.info("Finding data rows with NIK links...")
                all_rows = page.locator("table#vgt-table tbody tr").all()
                
                # Filter only rows that have NIK link
                data_rows = []
                for row in all_rows:
                    nik_link = row.locator('span[title="Lihat Detail Mitra"]')
                    if nik_link.count() > 0:
                        data_rows.append(row)
                
                self.stats['total'] = len(data_rows)
                
                if len(data_rows) == 0:
                    logger.error("✗ No data rows found in table!")
                    logger.info(f"Total rows found: {len(all_rows)}")
                    logger.info("Please ensure:")
                    logger.info("  1. You are logged in")
                    logger.info("  2. You are on the 'Seleksi Mitra' page")
                    logger.info("  3. The table has loaded completely")
                    logger.info("  4. There are actually data rows (not just headers)")
                    return
                
                logger.info(f"✓ Found {len(data_rows)} data rows (skipped {len(all_rows) - len(data_rows)} header rows)")
                
                # Detect total pages
                try:
                    page_info = page.locator('.footer__navigation__page-info__current-entry + span').inner_text()
                    total_pages = page_info.replace('dari', '').strip()
                    logger.info(f"✓ Detected total pages: {total_pages}")
                except:
                    total_pages = "unknown"
                    logger.info("⚠ Could not detect total pages")
                
                logger.info("\nStarting data extraction...\n")
                
                # Pagination loop
                current_page = 1
                has_next_page = True
                
                while has_next_page:
                    logger.info(f"\n{'='*60}")
                    logger.info(f"PROCESSING PAGE {current_page}")
                    logger.info(f"{'='*60}\n")
                    
                    # Process each data row on current page
                    for i, row in enumerate(data_rows):
                        self.process_row(row, i, page)
                    
                    # Increment pages counter
                    self.stats['pages_processed'] = current_page
                    
                    # Check if there's a next page button
                    try:
                        # Look for "Selanjutnya" button that's NOT disabled
                        # Based on HTML: <button class="footer__navigation__page-btn">Selanjutnya</button>
                        # Disabled buttons have class "disabled"
                        next_button = page.locator('button.footer__navigation__page-btn:has-text("Selanjutnya"):not(.disabled)')
                        
                        if next_button.count() > 0 and next_button.is_visible():
                            logger.info(f"\n✓ Page {current_page} completed. Moving to next page...")
                            next_button.click()
                            page.wait_for_timeout(3000)  # Wait for page to load
                            
                            # Re-fetch rows for new page
                            all_rows = page.locator("table#vgt-table tbody tr").all()
                            data_rows = []
                            for row in all_rows:
                                nik_link = row.locator('span[title="Lihat Detail Mitra"]')
                                if nik_link.count() > 0:
                                    data_rows.append(row)
                            
                            current_page += 1
                            logger.info(f"✓ Found {len(data_rows)} rows on page {current_page}")
                        else:
                            logger.info(f"\n✓ No more pages. Completed {current_page} page(s).")
                            has_next_page = False
                    except Exception as e:
                        logger.info(f"\n✓ Reached last page or pagination error: {str(e)}")
                        has_next_page = False

                
                logger.info("\n✓ All rows processed")
                
            except Exception as e:
                logger.error(f"✗ Fatal error: {str(e)}", exc_info=True)
                return
        
        # Save results
        if self.data_list:
            self.save_to_excel()
            self.save_to_csv()
        
        # Print summary
        self.print_summary()
        logger.info(f"\n✓ Scraping completed! Check {log_filename} for details.")

if __name__ == "__main__":
    scraper = MitraScraper()
    scraper.run()

