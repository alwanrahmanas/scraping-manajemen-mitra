import os
import sys
import logging
import requests
import csv
import re
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ijazah_parser import IjazahParser

# Setup logging
log_filename = f"scraper_test_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

class MitraScraperTest:
    def __init__(self, max_rows=5):
        # Create output folder with timestamp for versioning
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.output_folder = f"output_test_{timestamp}"
        self.base_download_dir = os.path.join(self.output_folder, "downloads")
        
        self.data_list = []
        self.max_rows = max_rows  # Limit untuk testing
        self.stats = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'ktp_downloaded': 0,
            'ijazah_downloaded': 0,
            'ijazah_parsed': 0,
            'pages_processed': 0
        }
        
        # Create output and downloads directory
        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)
            logger.info(f"Created output folder: {self.output_folder}")
        
        if not os.path.exists(self.base_download_dir):
            os.makedirs(self.base_download_dir)
            logger.info(f"Created downloads directory: {self.base_download_dir}")
        
        # Initialize Ijazah Parser (optional, akan skip jika API key tidak ada)
        self.ijazah_parser = None
        try:
            self.ijazah_parser = IjazahParser()
            logger.info("✓ IjazahParser initialized - Ijazah akan di-parse otomatis")
        except ValueError as e:
            logger.warning(f"⚠ IjazahParser tidak aktif: {e}")
            logger.warning("⚠ Ijazah akan didownload tapi tidak di-parse")
        except Exception as e:
            logger.warning(f"⚠ Error initializing IjazahParser: {e}")
            logger.warning("⚠ Ijazah akan didownload tapi tidak di-parse")

    def download_image(self, url, folder, filename):
        """Download image from URL with detailed logging"""
        if not url:
            logger.warning(f"No URL provided for {filename}")
            return None
        
        try:
            logger.info(f"Downloading {filename} from {url[:100]}...")
            response = requests.get(url, timeout=15)
            
            if response.status_code == 200:
                path = os.path.join(folder, filename)
                with open(path, 'wb') as f:
                    f.write(response.content)
                
                file_size = len(response.content) / 1024  # KB
                logger.info(f"✓ Downloaded {filename} ({file_size:.2f} KB) -> {path}")
                return path
            else:
                logger.error(f"✗ Failed to download {filename}: HTTP {response.status_code}")
                return None
                
        except Exception as e:
            logger.error(f"✗ Error downloading {filename}: {str(e)}")
            return None

    def extract_bank_info(self, page):
        """Extract bank information from Rekening tab"""
        logger.info("Extracting bank information...")
        
        nama_bank = "N/A"
        no_rekening = "N/A"
        nama_pemilik = "N/A"

        try:
            # Strategy 1: Direct extraction from form-control-plaintext
            try:
                bank_container = page.locator('label:has-text("Nama Bank") + div.form-control-plaintext')
                if bank_container.count() > 0:
                    nama_bank = bank_container.inner_text().strip()
                    logger.info(f"Found Nama Bank: {nama_bank}")
            except Exception as e:
                logger.debug(f"Strategy 1 failed for Nama Bank: {e}")
            
            try:
                rek_container = page.locator('label:has-text("Nomor Rekening") + div.form-control-plaintext')
                if rek_container.count() > 0:
                    no_rekening = rek_container.inner_text().strip()
                    logger.info(f"Found Nomor Rekening: {no_rekening}")
            except Exception as e:
                logger.debug(f"Strategy 1 failed for Nomor Rekening: {e}")
            
            try:
                owner_container = page.locator('label:has-text("Nama Pemilik Rekening") + div.form-control-plaintext')
                if owner_container.count() > 0:
                    nama_pemilik = owner_container.inner_text().strip()
                    logger.info(f"Found Nama Pemilik: {nama_pemilik}")
            except Exception as e:
                logger.debug(f"Strategy 1 failed for Nama Pemilik: {e}")
            
            # Fallback: Parse from text dump if any field is still N/A
            if nama_bank == "N/A" or no_rekening == "N/A" or nama_pemilik == "N/A":
                logger.info("Using fallback text parsing...")
                modal_text = page.locator(".v--modal-box").inner_text()
                lines = modal_text.split('\n')
                
                for i, line in enumerate(lines):
                    line = line.strip()
                    
                    if nama_bank == "N/A" and "Nama Bank" in line and i + 1 < len(lines):
                        potential_bank = lines[i + 1].strip()
                        if potential_bank and ("BANK" in potential_bank.upper() or potential_bank.startswith("(")):
                            nama_bank = potential_bank
                            logger.info(f"Fallback found Nama Bank: {nama_bank}")
                    
                    if no_rekening == "N/A" and "Nomor Rekening" in line and i + 1 < len(lines):
                        potential_rek = lines[i + 1].strip()
                        if potential_rek and (potential_rek.replace(' ', '').isdigit() or len(potential_rek) > 8):
                            no_rekening = potential_rek
                            logger.info(f"Fallback found Nomor Rekening: {no_rekening}")
                    
                    if nama_pemilik == "N/A" and "Nama Pemilik" in line and i + 1 < len(lines):
                        potential_owner = lines[i + 1].strip()
                        if potential_owner and len(potential_owner) > 2 and not potential_owner.isdigit():
                            nama_pemilik = potential_owner
                            logger.info(f"Fallback found Nama Pemilik: {nama_pemilik}")
        
        except Exception as e:
            logger.error(f"Error extracting bank info: {str(e)}")
        
        # Clean Nomor Rekening (Keep only numbers)
        if no_rekening != "N/A":
            no_rekening = re.sub(r'[^0-9]', '', no_rekening)
        
        return nama_bank, no_rekening, nama_pemilik

    def process_row(self, row, index, page):
        """Process a single table row"""
        try:
            # Find NIK link
            nik_link = row.locator('span[title="Lihat Detail Mitra"]')
            if not nik_link.count():
                logger.debug(f"Row {index}: No NIK link found, skipping")
                return False
            
            nik_text = nik_link.inner_text().strip()
            logger.info(f"\n{'='*60}")
            logger.info(f"Processing Row {index + 1}: NIK {nik_text}")
            logger.info(f"{'='*60}")
            
            # Create user directory
            user_download_dir = os.path.join(self.base_download_dir, nik_text)
            if not os.path.exists(user_download_dir):
                os.makedirs(user_download_dir)
                logger.info(f"Created directory: {user_download_dir}")
            
            # Click NIK to open popup
            logger.info("Opening detail popup...")
            nik_link.click()
            
            # Wait for modal with better error handling
            try:
                page.wait_for_selector("text=Detail Informasi Mitra", timeout=10000)
                logger.info("✓ Popup opened")
            except PlaywrightTimeoutError:
                logger.error("Timeout waiting for popup - trying to continue anyway")
                page.wait_for_timeout(2000)
            
            # === Tab 1: File Administrasi ===
            logger.info("\n--- Processing File Administrasi ---")
            try:
                # Try multiple selectors for File Administrasi tab
                file_admin_clicked = False
                selectors = [
                    '.nav-link:has-text("File Administrasi")',
                    '[role="tab"]:has-text("File Administrasi")',
                    'a:has-text("File Administrasi")'
                ]
                
                for selector in selectors:
                    try:
                        page.locator(selector).first.click(timeout=5000)
                        logger.info(f"✓ Clicked File Administrasi tab using: {selector}")
                        file_admin_clicked = True
                        break
                    except Exception:
                        continue
                
                if not file_admin_clicked:
                    logger.warning("Could not click File Administrasi tab - may already be active")
                    
            except Exception as e:
                logger.warning(f"Error clicking File Administrasi tab: {e}")
            
            page.wait_for_timeout(1500)
            
            ktp_path = None
            ijazah_path = None
            ijazah_data = None
            
            # Find images - gunakan selector berdasarkan label dan URL
            try:
                # Cari link KTP (foto_ktp/)
                ktp_links = page.locator('a[href*="foto_ktp/"]').all()
                if ktp_links:
                    ktp_href = ktp_links[0].get_attribute("href")
                    logger.info(f"Found KTP link: {ktp_href[:100]}...")
                    ktp_path = self.download_image(ktp_href, user_download_dir, "ktp.jpg")
                    if ktp_path:
                        self.stats['ktp_downloaded'] += 1
                else:
                    logger.warning("No KTP link found")
            except Exception as e:
                logger.error(f"Error finding KTP link: {e}")
            
            try:
                # Cari link Ijazah (ijazah/)
                ijazah_links = page.locator('a[href*="ijazah/"]').all()
                if ijazah_links:
                    ijazah_href = ijazah_links[0].get_attribute("href")
                    logger.info(f"Found Ijazah link: {ijazah_href[:100]}...")
                    ijazah_path = self.download_image(ijazah_href, user_download_dir, "ijazah.jpg")
                    if ijazah_path:
                        self.stats['ijazah_downloaded'] += 1
                        
                        # Parse ijazah jika parser tersedia
                        if self.ijazah_parser:
                            logger.info("Parsing ijazah dengan OpenAI Vision API...")
                            try:
                                ijazah_data = self.ijazah_parser.parse_ijazah(ijazah_path)
                                self.stats['ijazah_parsed'] += 1
                                logger.info(f"✓ Ijazah parsed successfully")
                            except Exception as e:
                                logger.error(f"✗ Error parsing ijazah: {e}")
                                ijazah_data = self.ijazah_parser._empty_result()
                        else:
                            ijazah_data = None
                else:
                    logger.warning("No Ijazah link found")
                    ijazah_data = None
            except Exception as e:
                logger.error(f"Error finding Ijazah link: {e}")
                ijazah_data = None
            
            # === Tab 2: Rekening ===
            logger.info("\n--- Processing Rekening ---")
            try:
                # Try multiple selectors for Rekening tab
                rekening_clicked = False
                selectors = [
                    '.nav-link:has-text("Rekening")',
                    '[role="tab"]:has-text("Rekening")',
                    'a:has-text("Rekening")'
                ]
                
                for selector in selectors:
                    try:
                        page.locator(selector).first.click(timeout=5000)
                        logger.info(f"✓ Clicked Rekening tab using: {selector}")
                        rekening_clicked = True
                        break
                    except Exception:
                        continue
                
                if not rekening_clicked:
                    logger.error("Failed to click Rekening tab with all selectors")
                    
            except Exception as e:
                logger.error(f"Error clicking Rekening tab: {e}")
            
            page.wait_for_timeout(1500)
            
            nama_bank, no_rekening, nama_pemilik = self.extract_bank_info(page)
            
            # Store data
            row_data = {
                "NIK": nik_text,
                "Nama Bank": nama_bank,
                "Nomor Rekening": no_rekening,
                "Nama Pemilik": nama_pemilik,
                "Path KTP": ktp_path if ktp_path else "Not Downloaded",
                "Path Ijazah": ijazah_path if ijazah_path else "Not Downloaded",
                "Status": "Success"
            }
            
            # Tambahkan data parsing ijazah jika tersedia
            if ijazah_data:
                row_data.update({
                    "Ijazah_Jenis": ijazah_data.get("jenis_ijazah", "N/A"),
                    "Ijazah_Nama": ijazah_data.get("nama", "N/A"),
                    "Ijazah_Gelar": ijazah_data.get("gelar", "N/A"),
                    "Ijazah_Nama_Gelar": ijazah_data.get("nama_gelar", "N/A"),
                    "Ijazah_NIM": ijazah_data.get("nim", "N/A"),
                    "Ijazah_Program_Studi": ijazah_data.get("program_studi", "N/A"),
                    "Ijazah_Fakultas": ijazah_data.get("fakultas", "N/A"),
                    "Ijazah_Universitas": ijazah_data.get("universitas", "N/A"),
                    "Ijazah_Tanggal": ijazah_data.get("tanggal_ijazah", "N/A")
                })
            else:
                row_data.update({
                    "Ijazah_Jenis": "N/A",
                    "Ijazah_Nama": "N/A",
                    "Ijazah_Gelar": "N/A",
                    "Ijazah_Nama_Gelar": "N/A",
                    "Ijazah_NIM": "N/A",
                    "Ijazah_Program_Studi": "N/A",
                    "Ijazah_Fakultas": "N/A",
                    "Ijazah_Universitas": "N/A",
                    "Ijazah_Tanggal": "N/A"
                })
            
            self.data_list.append(row_data)
            
            logger.info(f"\n✓ Successfully processed NIK {nik_text}")
            logger.info(f"  Bank: {nama_bank}")
            logger.info(f"  Rekening: {no_rekening}")
            logger.info(f"  Pemilik: {nama_pemilik}")
            if ijazah_data:
                logger.info(f"  Ijazah Nama: {ijazah_data.get('nama', 'N/A')}")
                logger.info(f"  Ijazah Gelar: {ijazah_data.get('gelar', 'N/A')}")
                logger.info(f"  Universitas: {ijazah_data.get('universitas', 'N/A')}")
            
            # Close modal with multiple attempts
            try:
                page.keyboard.press("Escape")
                page.wait_for_timeout(500)
            except Exception as e:
                logger.warning(f"Error closing modal with Escape: {e}")
                # Try clicking close button
                try:
                    page.locator('button.close, .modal-close, [aria-label="Close"]').first.click(timeout=2000)
                    page.wait_for_timeout(500)
                except Exception:
                    logger.warning("Could not close modal - continuing anyway")
            
            self.stats['success'] += 1
            return True
            
        except Exception as e:
            logger.error(f"✗ Error processing row {index}: {str(e)}", exc_info=True)
            self.stats['failed'] += 1
            
            # Try to close modal and recover
            try:
                page.keyboard.press("Escape")
                page.wait_for_timeout(500)
            except Exception:
                try:
                    page.locator('button.close, .modal-close, [aria-label="Close"]').first.click(timeout=2000)
                    page.wait_for_timeout(500)
                except Exception:
                    pass
            
            # Store failed entry
            self.data_list.append({
                "NIK": nik_text if 'nik_text' in locals() else "Unknown",
                "Nama Bank": "N/A",
                "Nomor Rekening": "N/A",
                "Nama Pemilik": "N/A",
                "Path KTP": "Failed",
                "Path Ijazah": "Failed",
                "Ijazah_Jenis": "N/A",
                "Ijazah_Nama": "N/A",
                "Ijazah_Gelar": "N/A",
                "Ijazah_Nama_Gelar": "N/A",
                "Ijazah_NIM": "N/A",
                "Ijazah_Program_Studi": "N/A",
                "Ijazah_Fakultas": "N/A",
                "Ijazah_Universitas": "N/A",
                "Ijazah_Tanggal": "N/A",
                "Status": f"Failed: {str(e)[:100]}"
            })
            
            return False

    def save_to_excel(self, filename="mitra_data_test.xlsx"):
        """Save data to Excel with formatting"""
        filepath = os.path.join(self.output_folder, filename)
        logger.info(f"\nSaving data to Excel: {filepath}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Mitra Test"
        
        # Headers - Kolom utama di depan
        headers = [
            "NIK", 
            "Nama Lengkap (dengan Gelar)",
            "Nomor Rekening",
            "Nama Bank",
            "Nama Pemilik Rekening",
            "Jenis Ijazah",
            "Gelar",
            "NIM",
            "Program Studi",
            "Fakultas",
            "Universitas",
            "Tanggal Ijazah",
            "Path KTP",
            "Path Ijazah",
            "Status"
        ]
        ws.append(headers)
        
        # Header formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data rows
        for row_data in self.data_list:
            ws.append([
                row_data.get("NIK", ""),
                row_data.get("Ijazah_Nama_Gelar", "N/A"),
                row_data.get("Nomor Rekening", ""),
                row_data.get("Nama Bank", ""),
                row_data.get("Nama Pemilik", ""),
                row_data.get("Ijazah_Jenis", ""),
                row_data.get("Ijazah_Gelar", ""),
                row_data.get("Ijazah_NIM", ""),
                row_data.get("Ijazah_Program_Studi", ""),
                row_data.get("Ijazah_Fakultas", ""),
                row_data.get("Ijazah_Universitas", ""),
                row_data.get("Ijazah_Tanggal", ""),
                row_data.get("Path KTP", ""),
                row_data.get("Path Ijazah", ""),
                row_data.get("Status", "")
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        
        wb.save(filepath)
        logger.info(f"✓ Excel file saved: {filepath}")

    def save_to_csv(self, filename="mitra_data_test.csv"):
        """Save data to CSV"""
        filepath = os.path.join(self.output_folder, filename)
        logger.info(f"Saving CSV backup: {filepath}")
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            fieldnames = [
                "NIK", "Ijazah_Nama_Gelar", "Nomor Rekening",
                "Nama Bank", "Nama Pemilik", 
                "Ijazah_Jenis", "Ijazah_Nama", "Ijazah_Gelar", "Ijazah_NIM",
                "Ijazah_Program_Studi", "Ijazah_Fakultas", "Ijazah_Universitas", "Ijazah_Tanggal",
                "Path KTP", "Path Ijazah", "Status"
            ]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(self.data_list)
        
        logger.info(f"✓ CSV backup saved: {filepath}")

    def print_summary(self):
        """Print scraping summary"""
        logger.info(f"\n{'='*60}")
        logger.info("TESTING SUMMARY")
        logger.info(f"{'='*60}")
        logger.info(f"Max rows (limit): {self.max_rows}")
        logger.info(f"Total rows processed: {self.stats['total']}")
        logger.info(f"✓ Successful: {self.stats['success']}")
        logger.info(f"✗ Failed: {self.stats['failed']}")
        logger.info(f"📷 KTP downloaded: {self.stats['ktp_downloaded']}")
        logger.info(f"📷 Ijazah downloaded: {self.stats['ijazah_downloaded']}")
        logger.info(f"📝 Ijazah parsed: {self.stats['ijazah_parsed']}")
        logger.info(f"{'='*60}")

    def run(self):
        """Main scraping process - LIMITED TO max_rows"""
        logger.info("="*60)
        logger.info(f"MITRA BPS SCRAPER - TESTING MODE (Max {self.max_rows} rows)")
        logger.info("="*60)
        logger.info(f"Log file: {log_filename}")
        
        with sync_playwright() as p:
            try:
                logger.info("\nConnecting to Chrome (port 9222)...")
                browser = p.chromium.connect_over_cdp("http://localhost:9222")
                context = browser.contexts[0]
                
                # Cari tab yang benar (skip DevTools dan fs-storage)
                page = None
                for p_page in context.pages:
                    url = p_page.url
                    
                    # Skip DevTools dan foto tabs
                    if "devtools://" in url or "fs-storage" in url:
                        continue
                    
                    # Gunakan tab pertama yang valid
                    page = p_page
                    break
                
                if not page:
                    logger.error("✗ No suitable tab found! Please open Seleksi Mitra page in Chrome.")
                    return
                
                logger.info(f"✓ Connected to: {page.title()}")
                logger.info(f"✓ URL: {page.url}")
                
                # Wait for table
                logger.info("\nWaiting for table to load...")
                page.wait_for_timeout(2000)
                
                # Get all rows
                logger.info("Finding data rows with NIK links...")
                all_rows = page.locator("table#vgt-table tbody tr").all()
                
                # Filter rows with NIK link
                data_rows = []
                for row in all_rows:
                    nik_link = row.locator('span[title="Lihat Detail Mitra"]')
                    if nik_link.count() > 0:
                        data_rows.append(row)
                        # LIMIT: Stop jika sudah mencapai max_rows
                        if len(data_rows) >= self.max_rows:
                            break
                
                self.stats['total'] = len(data_rows)
                
                if len(data_rows) == 0:
                    logger.error("✗ No data rows found in table!")
                    return
                
                logger.info(f"✓ Found {len(data_rows)} data rows (limited to {self.max_rows})")
                logger.info("\nStarting data extraction...\n")
                
                # Process rows (limited)
                for i, row in enumerate(data_rows):
                    self.process_row(row, i, page)
                    
                    # Progress indicator
                    logger.info(f"\n--- Progress: {i+1}/{len(data_rows)} ---\n")
                    
                    # Small delay between rows to avoid overwhelming the system
                    page.wait_for_timeout(500)
                
                logger.info("\n✓ All rows processed")
                
            except Exception as e:
                logger.error(f"✗ Fatal error: {str(e)}", exc_info=True)
                return
        
        # Save results
        if self.data_list:
            self.save_to_excel()
            self.save_to_csv()
        else:
            logger.warning("⚠ No data collected - skipping file save")
        
        # Print summary
        self.print_summary()
        logger.info(f"\n✓ Testing completed! Check {log_filename} for details.")

if __name__ == "__main__":
    # Testing dengan 3 orang pertama
    scraper = MitraScraperTest(max_rows=3)
    scraper.run()