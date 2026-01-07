import os
import sys
import logging
import requests
import csv
import re
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ijazah_parser import IjazahParser

# Setup logging
log_filename = f"scraper_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

class MitraScraper:
    def __init__(self):
        # Create output folder with timestamp for versioning
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.output_folder = f"output_{timestamp}"
        self.base_download_dir = os.path.join(self.output_folder, "downloads")
        
        self.data_list = []
        self.stats = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'ktp_downloaded': 0,
            'ijazah_downloaded': 0,
            'ijazah_parsed': 0,
            'pages_processed': 0
        }
        
        # Create output and downloads directory
        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)
            logger.info(f"Created output folder: {self.output_folder}")
        
        if not os.path.exists(self.base_download_dir):
            os.makedirs(self.base_download_dir)
            logger.info(f"Created downloads directory: {self.base_download_dir}")
        
        # Initialize Ijazah Parser (optional, akan skip jika API key tidak ada)
        self.ijazah_parser = None
        try:
            self.ijazah_parser = IjazahParser()
            logger.info("✓ IjazahParser initialized - Ijazah akan di-parse otomatis")
        except ValueError as e:
            logger.warning(f"⚠ IjazahParser tidak aktif: {e}")
            logger.warning("⚠ Ijazah akan didownload tapi tidak di-parse")
        except Exception as e:
            logger.warning(f"⚠ Error initializing IjazahParser: {e}")
            logger.warning("⚠ Ijazah akan didownload tapi tidak di-parse")

    def download_image(self, url, folder, filename):
        """Download image from URL with detailed logging"""
        if not url:
            logger.warning(f"No URL provided for {filename}")
            return None
        
        try:
            logger.info(f"Downloading {filename} from {url[:100]}...")
            response = requests.get(url, timeout=15)
            
            if response.status_code == 200:
                path = os.path.join(folder, filename)
                with open(path, 'wb') as f:
                    f.write(response.content)
                
                file_size = len(response.content) / 1024  # KB
                logger.info(f"✓ Downloaded {filename} ({file_size:.2f} KB) -> {path}")
                return path
            else:
                logger.error(f"✗ Failed to download {filename}: HTTP {response.status_code}")
                return None
                
        except Exception as e:
            logger.error(f"✗ Error downloading {filename}: {str(e)}")
            return None

    def extract_bank_info(self, page):
        """Extract bank information from Rekening tab"""
        logger.info("Extracting bank information...")
        
        nama_bank = "N/A"
        no_rekening = "N/A"
        nama_pemilik = "N/A"

        try:
            # Strategy 1: Direct extraction from form-control-plaintext
            try:
                bank_container = page.locator('label:has-text("Nama Bank") + div.form-control-plaintext')
                if bank_container.count() > 0:
                    nama_bank = bank_container.inner_text().strip()
                    logger.info(f"Found Nama Bank: {nama_bank}")
            except Exception as e:
                logger.debug(f"Strategy 1 failed for Nama Bank: {e}")
            
            try:
                rek_container = page.locator('label:has-text("Nomor Rekening") + div.form-control-plaintext')
                if rek_container.count() > 0:
                    no_rekening = rek_container.inner_text().strip()
                    logger.info(f"Found Nomor Rekening: {no_rekening}")
            except Exception as e:
                logger.debug(f"Strategy 1 failed for Nomor Rekening: {e}")
            
            try:
                owner_container = page.locator('label:has-text("Nama Pemilik Rekening") + div.form-control-plaintext')
                if owner_container.count() > 0:
                    nama_pemilik = owner_container.inner_text().strip()
                    logger.info(f"Found Nama Pemilik: {nama_pemilik}")
            except Exception as e:
                logger.debug(f"Strategy 1 failed for Nama Pemilik: {e}")
            
            # Fallback: Parse from text dump if any field is still N/A
            if nama_bank == "N/A" or no_rekening == "N/A" or nama_pemilik == "N/A":
                logger.info("Using fallback text parsing...")
                modal_text = page.locator(".v--modal-box").inner_text()
                lines = modal_text.split('\n')
                
                for i, line in enumerate(lines):
                    line = line.strip()
                    
                    if nama_bank == "N/A" and "Nama Bank" in line and i + 1 < len(lines):
                        potential_bank = lines[i + 1].strip()
                        if potential_bank and ("BANK" in potential_bank.upper() or potential_bank.startswith("(")):
                            nama_bank = potential_bank
                            logger.info(f"Fallback found Nama Bank: {nama_bank}")
                    
                    if no_rekening == "N/A" and "Nomor Rekening" in line and i + 1 < len(lines):
                        potential_rek = lines[i + 1].strip()
                        if potential_rek and (potential_rek.replace(' ', '').isdigit() or len(potential_rek) > 8):
                            no_rekening = potential_rek
                            logger.info(f"Fallback found Nomor Rekening: {no_rekening}")
                    
                    if nama_pemilik == "N/A" and "Nama Pemilik" in line and i + 1 < len(lines):
                        potential_owner = lines[i + 1].strip()
                        if potential_owner and len(potential_owner) > 2 and not potential_owner.isdigit():
                            nama_pemilik = potential_owner
                            logger.info(f"Fallback found Nama Pemilik: {nama_pemilik}")
        
        except Exception as e:
            logger.error(f"Error extracting bank info: {str(e)}")
        
        # Clean Nomor Rekening (Keep only numbers)
        if no_rekening != "N/A":
            no_rekening = re.sub(r'[^0-9]', '', no_rekening)
        
        return nama_bank, no_rekening, nama_pemilik

    def process_row(self, row, index, page):
        """Process a single table row"""
        try:
            # Find NIK link
            nik_link = row.locator('span[title="Lihat Detail Mitra"]')
            if not nik_link.count():
                logger.debug(f"Row {index}: No NIK link found, skipping")
                return False
            
            nik_text = nik_link.inner_text().strip()
            logger.info(f"\n{'='*60}")
            logger.info(f"Processing Row {index + 1}: NIK {nik_text}")
            logger.info(f"{'='*60}")
            
            # Create user directory
            user_download_dir = os.path.join(self.base_download_dir, nik_text)
            if not os.path.exists(user_download_dir):
                os.makedirs(user_download_dir)
                logger.info(f"Created directory: {user_download_dir}")
            
            # Click NIK to open popup
            logger.info("Opening detail popup...")
            nik_link.click()
            
            # Wait for modal with better error handling
            try:
                page.wait_for_selector("text=Detail Informasi Mitra", timeout=10000)
                logger.info("✓ Popup opened")
            except PlaywrightTimeoutError:
                logger.error("Timeout waiting for popup - trying to continue anyway")
                page.wait_for_timeout(2000)
            
            # === Tab 1: File Administrasi ===
            logger.info("\n--- Processing File Administrasi ---")
            try:
                # Try multiple selectors for File Administrasi tab
                file_admin_clicked = False
                selectors = [
                    '.nav-link:has-text("File Administrasi")',
                    '[role="tab"]:has-text("File Administrasi")',
                    'a:has-text("File Administrasi")'
                ]
                
                for selector in selectors:
                    try:
                        page.locator(selector).first.click(timeout=5000)
                        logger.info(f"✓ Clicked File Administrasi tab using: {selector}")
                        file_admin_clicked = True
                        break
                    except Exception:
                        continue
                
                if not file_admin_clicked:
                    logger.warning("Could not click File Administrasi tab - may already be active")
                    
            except Exception as e:
                logger.warning(f"Error clicking File Administrasi tab: {e}")
            
            page.wait_for_timeout(1500)
            
            ktp_path = None
            ijazah_path = None
            ijazah_data = None
            
            # Find images - gunakan selector berdasarkan label dan URL
            try:
                # Cari link KTP (foto_ktp/)
                ktp_links = page.locator('a[href*="foto_ktp/"]').all()
                if ktp_links:
                    ktp_href = ktp_links[0].get_attribute("href")
                    logger.info(f"Found KTP link: {ktp_href[:100]}...")
                    ktp_path = self.download_image(ktp_href, user_download_dir, "ktp.jpg")
                    if ktp_path:
                        self.stats['ktp_downloaded'] += 1
                else:
                    logger.warning("No KTP link found")
            except Exception as e:
                logger.error(f"Error finding KTP link: {e}")
            
            try:
                # Cari link Ijazah (ijazah/)
                ijazah_links = page.locator('a[href*="ijazah/"]').all()
                if ijazah_links:
                    ijazah_href = ijazah_links[0].get_attribute("href")
                    logger.info(f"Found Ijazah link: {ijazah_href[:100]}...")
                    ijazah_path = self.download_image(ijazah_href, user_download_dir, "ijazah.jpg")
                    if ijazah_path:
                        self.stats['ijazah_downloaded'] += 1
                        
                        # Parse ijazah jika parser tersedia
                        if self.ijazah_parser:
                            logger.info("Parsing ijazah dengan OpenAI Vision API...")
                            try:
                                ijazah_data = self.ijazah_parser.parse_ijazah(ijazah_path)
                                self.stats['ijazah_parsed'] += 1
                                logger.info(f"✓ Ijazah parsed successfully")
                            except Exception as e:
                                logger.error(f"✗ Error parsing ijazah: {e}")
                                ijazah_data = self.ijazah_parser._empty_result()
                        else:
                            ijazah_data = None
                else:
                    logger.warning("No Ijazah link found")
                    ijazah_data = None
            except Exception as e:
                logger.error(f"Error finding Ijazah link: {e}")
                ijazah_data = None
            
            # === Tab 2: Rekening ===
            logger.info("\n--- Processing Rekening ---")
            try:
                # Try multiple selectors for Rekening tab
                rekening_clicked = False
                selectors = [
                    '.nav-link:has-text("Rekening")',
                    '[role="tab"]:has-text("Rekening")',
                    'a:has-text("Rekening")'
                ]
                
                for selector in selectors:
                    try:
                        page.locator(selector).first.click(timeout=10000)
                        logger.info(f"✓ Clicked Rekening tab using: {selector}")
                        rekening_clicked = True
                        break
                    except Exception:
                        continue
                
                if not rekening_clicked:
                    logger.error("Failed to click Rekening tab with all selectors")
                    
            except Exception as e:
                logger.error(f"Error clicking Rekening tab: {e}")
            
            # Wait for Rekening tab content to fully load (prevent race condition)
            try:
                page.wait_for_selector('label:has-text("Nama Bank")', state="visible", timeout=8000)
                logger.info("✓ Rekening tab content loaded")
                page.wait_for_timeout(800)  # Extra buffer for dynamic content
            except PlaywrightTimeoutError:
                logger.warning("⚠ Rekening content load timeout - continuing anyway")
                page.wait_for_timeout(2000)
            
            nama_bank, no_rekening, nama_pemilik = self.extract_bank_info(page)
            
            # Validate scraped data to detect potential mismatch
            has_mismatch = False
            if no_rekening != "N/A" and not no_rekening.replace('-', '').replace(' ', '').isdigit():
                has_mismatch = True
                logger.error(f"⚠ POTENTIAL MISMATCH DETECTED for NIK {nik_text}!")
                logger.error(f"  Nomor Rekening contains non-numeric: '{no_rekening}'")
                logger.error(f"  Nama Pemilik: '{nama_pemilik}'")
                logger.error(f"  This data may be incorrect - please verify manually!")
            
            # Store data
            row_data = {
                "NIK": nik_text,
                "Nama Bank": nama_bank,
                "Nomor Rekening": no_rekening,
                "Nama Pemilik": nama_pemilik,
                "Path KTP": ktp_path if ktp_path else "Not Downloaded",
                "Path Ijazah": ijazah_path if ijazah_path else "Not Downloaded",
                "Status": "Success",
                "_has_mismatch": has_mismatch  # Internal flag for Excel highlighting
            }
            
            # Tambahkan data parsing ijazah jika tersedia
            if ijazah_data:
                row_data.update({
                    "Ijazah_Jenis": ijazah_data.get("jenis_ijazah", "N/A"),
                    "Ijazah_Nama": ijazah_data.get("nama", "N/A"),
                    "Ijazah_Gelar": ijazah_data.get("gelar", "N/A"),
                    "Ijazah_Nama_Gelar": ijazah_data.get("nama_gelar", "N/A"),
                    "Ijazah_NIM": ijazah_data.get("nim", "N/A"),
                    "Ijazah_Program_Studi": ijazah_data.get("program_studi", "N/A"),
                    "Ijazah_Fakultas": ijazah_data.get("fakultas", "N/A"),
                    "Ijazah_Universitas": ijazah_data.get("universitas", "N/A"),
                    "Ijazah_Tanggal": ijazah_data.get("tanggal_ijazah", "N/A")
                })
            else:
                row_data.update({
                    "Ijazah_Jenis": "N/A",
                    "Ijazah_Nama": "N/A",
                    "Ijazah_Gelar": "N/A",
                    "Ijazah_Nama_Gelar": "N/A",
                    "Ijazah_NIM": "N/A",
                    "Ijazah_Program_Studi": "N/A",
                    "Ijazah_Fakultas": "N/A",
                    "Ijazah_Universitas": "N/A",
                    "Ijazah_Tanggal": "N/A"
                })
            
            self.data_list.append(row_data)
            
            logger.info(f"\n✓ Successfully processed NIK {nik_text}")
            logger.info(f"  Bank: {nama_bank}")
            logger.info(f"  Rekening: {no_rekening}")
            logger.info(f"  Pemilik: {nama_pemilik}")
            if ijazah_data:
                logger.info(f"  Ijazah Nama: {ijazah_data.get('nama', 'N/A')}")
                logger.info(f"  Ijazah Gelar: {ijazah_data.get('gelar', 'N/A')}")
                logger.info(f"  Universitas: {ijazah_data.get('universitas', 'N/A')}")
            
            # Close modal with multiple attempts and verification
            try:
                page.keyboard.press("Escape")
                page.wait_for_timeout(500)
                # Verify modal is actually closed
                try:
                    page.wait_for_selector(".v--modal-box", state="hidden", timeout=3000)
                    logger.info("✓ Modal closed successfully")
                except PlaywrightTimeoutError:
                    logger.warning("⚠ Modal may still be visible after Escape")
            except Exception as e:
                logger.warning(f"Error closing modal with Escape: {e}")
                # Try clicking close button
                try:
                    page.locator('button.close, .modal-close, [aria-label="Close"]').first.click(timeout=2000)
                    page.wait_for_timeout(500)
                    page.wait_for_selector(".v--modal-box", state="hidden", timeout=3000)
                    logger.info("✓ Modal closed via button")
                except Exception:
                    logger.warning("⚠ Could not verify modal closure - continuing anyway")
            
            self.stats['success'] += 1
            return True
            
        except Exception as e:
            logger.error(f"✗ Error processing row {index}: {str(e)}", exc_info=True)
            self.stats['failed'] += 1
            
            # Try to close modal and recover
            try:
                page.keyboard.press("Escape")
                page.wait_for_timeout(500)
            except Exception:
                try:
                    page.locator('button.close, .modal-close, [aria-label="Close"]').first.click(timeout=2000)
                    page.wait_for_timeout(500)
                except Exception:
                    pass
            
            # Store failed entry
            self.data_list.append({
                "NIK": nik_text if 'nik_text' in locals() else "Unknown",
                "Nama Bank": "N/A",
                "Nomor Rekening": "N/A",
                "Nama Pemilik": "N/A",
                "Path KTP": "Failed",
                "Path Ijazah": "Failed",
                "Ijazah_Jenis": "N/A",
                "Ijazah_Nama": "N/A",
                "Ijazah_Gelar": "N/A",
                "Ijazah_Nama_Gelar": "N/A",
                "Ijazah_NIM": "N/A",
                "Ijazah_Program_Studi": "N/A",
                "Ijazah_Fakultas": "N/A",
                "Ijazah_Universitas": "N/A",
                "Ijazah_Tanggal": "N/A",
                "Status": f"Failed: {str(e)[:100]}"
            })
            
            return False

    def save_to_excel(self, filename="mitra_data.xlsx"):
        """Save data to Excel with formatting"""
        filepath = os.path.join(self.output_folder, filename)
        logger.info(f"\nSaving data to Excel: {filepath}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Mitra"
        
        # Headers - Kolom utama di depan
        headers = [
            "NIK", 
            "Nama Lengkap (dengan Gelar)",
            "Nomor Rekening",
            "Nama Bank",
            "Nama Pemilik Rekening",
            "Jenis Ijazah",
            "Gelar",
            "NIM",
            "Program Studi",
            "Fakultas",
            "Universitas",
            "Tanggal Ijazah",
            "Path KTP",
            "Path Ijazah",
            "Status"
        ]
        ws.append(headers)
        
        # Header formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data rows with conditional formatting
        mismatch_count = 0
        for row_idx, row_data in enumerate(self.data_list, start=2):  # Start at row 2 (after header)
            ws.append([
                row_data.get("NIK", ""),
                row_data.get("Ijazah_Nama_Gelar", "N/A"),
                row_data.get("Nomor Rekening", ""),
                row_data.get("Nama Bank", ""),
                row_data.get("Nama Pemilik", ""),
                row_data.get("Ijazah_Jenis", ""),
                row_data.get("Ijazah_Gelar", ""),
                row_data.get("Ijazah_NIM", ""),
                row_data.get("Ijazah_Program_Studi", ""),
                row_data.get("Ijazah_Fakultas", ""),
                row_data.get("Ijazah_Universitas", ""),
                row_data.get("Ijazah_Tanggal", ""),
                row_data.get("Path KTP", ""),
                row_data.get("Path Ijazah", ""),
                row_data.get("Status", "")
            ])
            
            # Highlight rows with potential mismatch
            if row_data.get("_has_mismatch", False):
                mismatch_count += 1
                # Red/orange fill for mismatch rows
                mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                mismatch_font = Font(color="9C0006", bold=True)
                
                for cell in ws[row_idx]:
                    cell.fill = mismatch_fill
                    # Highlight Nomor Rekening column specifically
                    if cell.column == 3:  # Column C (Nomor Rekening)
                        cell.font = mismatch_font
        
        logger.info(f"✓ Highlighted {mismatch_count} rows with potential mismatch")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        
        # Create Summary Sheet
        ws_summary = wb.create_sheet("Summary", 0)  # Insert at beginning
        ws_summary.append(["SCRAPING SUMMARY & DATA QUALITY REPORT"])
        ws_summary.append([])
        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["Total Rows Scraped", len(self.data_list)])
        ws_summary.append(["Rows with Potential Mismatch", mismatch_count])
        ws_summary.append(["Data Quality Rate", f"{((len(self.data_list) - mismatch_count) / len(self.data_list) * 100):.1f}%" if self.data_list else "N/A"])
        ws_summary.append([])
        ws_summary.append(["LEGEND:"])
        ws_summary.append(["🔴 Red/Pink Rows", "= Potential mismatch detected (Nomor Rekening contains non-numeric characters)"])
        ws_summary.append(["⚠️ Action Required", "= Please verify these rows manually"])
        ws_summary.append([])
        ws_summary.append(["Note:", "Mismatch detection helps identify data quality issues where account number may have been incorrectly scraped."])
        
        # Format summary sheet
        ws_summary['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws_summary['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws_summary['A1'].alignment = Alignment(horizontal='center')
        ws_summary.merge_cells('A1:B1')
        
        # Format metric headers
        for cell in ws_summary[3]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Highlight mismatch count if > 0
        if mismatch_count > 0:
            ws_summary['B5'].font = Font(bold=True, color="9C0006")
            ws_summary['B5'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Auto-adjust summary column widths
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 60
        
        wb.save(filepath)
        logger.info(f"✓ Excel file saved: {filepath}")
        if mismatch_count > 0:
            logger.info(f"⚠ {mismatch_count} rows highlighted in red - please verify manually!")

    def save_to_csv(self, filename="mitra_data.csv"):
        """Save data to CSV"""
        filepath = os.path.join(self.output_folder, filename)
        logger.info(f"Saving CSV backup: {filepath}")
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            fieldnames = [
                "NIK", "Ijazah_Nama_Gelar", "Nomor Rekening",
                "Nama Bank", "Nama Pemilik", 
                "Ijazah_Jenis", "Ijazah_Nama", "Ijazah_Gelar", "Ijazah_NIM",
                "Ijazah_Program_Studi", "Ijazah_Fakultas", "Ijazah_Universitas", "Ijazah_Tanggal",
                "Path KTP", "Path Ijazah", "Status"
            ]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(self.data_list)
        
        logger.info(f"✓ CSV backup saved: {filepath}")

    def print_summary(self):
        """Print scraping summary"""
        logger.info(f"\n{'='*60}")
        logger.info("SCRAPING SUMMARY")
        logger.info(f"{'='*60}")
        logger.info(f"Pages processed: {self.stats['pages_processed']}")
        logger.info(f"Total rows processed: {self.stats['total']}")
        logger.info(f"✓ Successful: {self.stats['success']}")
        logger.info(f"✗ Failed: {self.stats['failed']}")
        logger.info(f"📷 KTP downloaded: {self.stats['ktp_downloaded']}")
        logger.info(f"📷 Ijazah downloaded: {self.stats['ijazah_downloaded']}")
        logger.info(f"📝 Ijazah parsed: {self.stats['ijazah_parsed']}")
        logger.info(f"{'='*60}")

    def run(self):
        """Main scraping process"""
        logger.info("="*60)
        logger.info("MITRA BPS SCRAPER - STARTING")
        logger.info("="*60)
        logger.info(f"Log file: {log_filename}")
        
        with sync_playwright() as p:
            try:
                logger.info("\nConnecting to Chrome (port 9222)...")
                browser = p.chromium.connect_over_cdp("http://localhost:9222")
                context = browser.contexts[0]
                
                # Cari tab yang benar (skip DevTools dan fs-storage)
                page = None
                for p_page in context.pages:
                    url = p_page.url
                    
                    # Skip DevTools dan foto tabs
                    if "devtools://" in url or "fs-storage" in url:
                        continue
                    
                    # Gunakan tab pertama yang valid
                    page = p_page
                    break
                
                if not page:
                    logger.error("✗ No suitable tab found! Please open Seleksi Mitra page in Chrome.")
                    return
                
                logger.info(f"✓ Connected to: {page.title()}")
                logger.info(f"✓ URL: {page.url}")
                
                # Wait for table with multiple strategies
                logger.info("\nWaiting for table to load...")
                
                # Strategy 1: Wait for table to exist
                try:
                    page.wait_for_selector("table#vgt-table tbody tr", state="attached", timeout=5000)
                    logger.info("✓ Table found (attached)")
                except Exception:
                    logger.warning("Table not found with 'attached' state, trying alternative...")
                
                # Strategy 2: Wait for loading overlay to disappear
                try:
                    overlay = page.locator(".velmld-overlay")
                    if overlay.count() > 0:
                        logger.info("Waiting for loading overlay to disappear...")
                        page.wait_for_selector(".velmld-overlay", state="hidden", timeout=15000)
                        logger.info("✓ Loading overlay hidden")
                except Exception:
                    logger.info("No loading overlay found or already hidden")
                
                # Strategy 3: Just wait a bit for any animations
                page.wait_for_timeout(2000)
                
                # Get all rows
                logger.info("Finding data rows with NIK links...")
                all_rows = page.locator("table#vgt-table tbody tr").all()
                
                # Filter rows with NIK link
                data_rows = []
                for row in all_rows:
                    nik_link = row.locator('span[title="Lihat Detail Mitra"]')
                    if nik_link.count() > 0:
                        data_rows.append(row)
                
                self.stats['total'] = len(data_rows)
                
                if len(data_rows) == 0:
                    logger.error("✗ No data rows found in table!")
                    logger.info(f"Total rows found: {len(all_rows)}")
                    logger.info("Please ensure:")
                    logger.info("  1. You are logged in")
                    logger.info("  2. You are on the 'Seleksi Mitra' page")
                    logger.info("  3. The table has loaded completely")
                    logger.info("  4. There are actually data rows (not just headers)")
                    return
                
                logger.info(f"✓ Found {len(data_rows)} data rows (skipped {len(all_rows) - len(data_rows)} header rows)")
                
                # Detect total pages
                try:
                    page_info = page.locator('.footer__navigation__page-info__current-entry + span').inner_text()
                    total_pages = page_info.replace('dari', '').strip()
                    logger.info(f"✓ Detected total pages: {total_pages}")
                except Exception:
                    total_pages = "unknown"
                    logger.info("⚠ Could not detect total pages")
                
                logger.info("\nStarting data extraction...\n")
                
                # Pagination loop
                current_page = 1
                has_next_page = True
                
                while has_next_page:
                    logger.info(f"\n{'='*60}")
                    logger.info(f"PROCESSING PAGE {current_page}")
                    logger.info(f"{'='*60}\n")
                    
                    # Process each data row on current page
                    for i, row in enumerate(data_rows):
                        self.process_row(row, i, page)
                        
                        # Small delay between rows
                        page.wait_for_timeout(500)
                    
                    # Increment pages counter
                    self.stats['pages_processed'] = current_page
                    
                    # Check if there's a next page button
                    try:
                        # Look for "Selanjutnya" button that's NOT disabled
                        next_button = page.locator('button.footer__navigation__page-btn:has-text("Selanjutnya"):not(.disabled)')
                        
                        if next_button.count() > 0 and next_button.is_visible():
                            logger.info(f"\n✓ Page {current_page} completed. Moving to next page...")
                            next_button.click()
                            page.wait_for_timeout(3000)  # Wait for page to load
                            
                            # Wait for loading overlay to disappear
                            try:
                                overlay = page.locator(".velmld-overlay")
                                if overlay.count() > 0:
                                    page.wait_for_selector(".velmld-overlay", state="hidden", timeout=15000)
                            except Exception:
                                pass
                            
                            # Re-fetch rows for new page
                            all_rows = page.locator("table#vgt-table tbody tr").all()
                            data_rows = []
                            for row in all_rows:
                                nik_link = row.locator('span[title="Lihat Detail Mitra"]')
                                if nik_link.count() > 0:
                                    data_rows.append(row)
                            
                            current_page += 1
                            logger.info(f"✓ Found {len(data_rows)} rows on page {current_page}")
                        else:
                            logger.info(f"\n✓ No more pages. Completed {current_page} page(s).")
                            has_next_page = False
                    except Exception as e:
                        logger.info(f"\n✓ Reached last page or pagination error: {str(e)}")
                        has_next_page = False
                
                logger.info("\n✓ All rows processed")
                
            except Exception as e:
                logger.error(f"✗ Fatal error: {str(e)}", exc_info=True)
                return
        
        # Save results
        if self.data_list:
            self.save_to_excel()
            self.save_to_csv()
        else:
            logger.warning("⚠ No data collected - skipping file save")
        
        # Print summary
        self.print_summary()
        logger.info(f"\n✓ Scraping completed! Check {log_filename} for details.")

if __name__ == "__main__":
    scraper = MitraScraper()
    scraper.run()